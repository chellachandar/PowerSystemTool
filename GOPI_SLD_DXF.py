import streamlit as st
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle, Polygon, Arc
import matplotlib.patches as mpatches
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import os
import tempfile
import io
import math
import ezdxf

# --- DXF CANVAS SCRAPER ENGINE ---
# This mirrors Matplotlib output to DXF without touching your drawing logic
def export_ax_to_dxf(ax):
    doc = ezdxf.new('R2010')
    msp = doc.modelspace()
    
    # 1. Extract and map lines
    for line in ax.lines:
        xdata = line.get_xdata()
        ydata = line.get_ydata()
        color = line.get_color()
        dxf_color = 7 # Default Black/White
        if color == 'red': dxf_color = 1
        elif color == 'green': dxf_color = 3
        elif color == 'blue': dxf_color = 5
        
        for idx in range(len(xdata)-1):
            msp.add_line((xdata[idx], ydata[idx]), (xdata[idx+1], ydata[idx+1]), dxfattribs={'color': dxf_color})

    # 2. Extract and map geometry (Arcs, Rectangles, Polygons)
    for patch in ax.patches:
        ec = patch.get_edgecolor()
        dxf_color = 7
        # Convert RGBA tuple back to DXF Colors
        if ec[0] > 0.5 and ec[1] < 0.5 and ec[2] < 0.5: dxf_color = 1
        elif ec[1] > 0.5 and ec[0] < 0.5 and ec[2] < 0.5: dxf_color = 3
        elif ec[2] > 0.5 and ec[0] < 0.5 and ec[1] < 0.5: dxf_color = 5
        
        if isinstance(patch, mpatches.Rectangle):
            x, y = patch.get_xy()
            w, h = patch.get_width(), patch.get_height()
            msp.add_lwpolyline([(x, y), (x+w, y), (x+w, y+h), (x, y+h)], close=True, dxfattribs={'color': dxf_color})
        
        elif isinstance(patch, mpatches.Polygon):
            pts = patch.get_xy()
            msp.add_lwpolyline(pts, close=True, dxfattribs={'color': dxf_color})
        
        elif isinstance(patch, mpatches.Arc):
            w, h = patch.width, patch.height
            cx, cy = patch.center
            rad_angle = math.radians(patch.angle)
            
            mx = (w/2.0) * math.cos(rad_angle)
            my = (w/2.0) * math.sin(rad_angle)
            ratio = h / w if w > 0 else 1.0
            
            t1, t2 = patch.theta1, patch.theta2
            if ratio > 1.0: # ezdxf handles minor/major axis strictly
                mx, my = -my, mx
                ratio = 1.0 / ratio
                t1 -= 90
                t2 -= 90
                
            try:
                msp.add_ellipse((cx, cy), major_axis=(mx, my), ratio=ratio,
                                start_param=math.radians(t1), end_param=math.radians(t2),
                                dxfattribs={'color': dxf_color})
            except Exception:
                pass

    # 3. Extract and map Text Labels
    for txt in ax.texts:
        x, y = txt.get_position()
        text_str = txt.get_text()
        fs = txt.get_fontsize()
        ha = txt.get_ha()
        align = 'MIDDLE_CENTER'
        if ha == 'left': align = 'MIDDLE_LEFT'
        elif ha == 'right': align = 'MIDDLE_RIGHT'
        
        lines = text_str.split('\n')
        y_off = 0
        for l_str in lines:
            if l_str.strip():
                dtxt = msp.add_text(l_str, dxfattribs={'height': fs / 10.0, 'color': 7})
                dtxt.set_placement((x, y - y_off), align=align)
            y_off += (fs / 10.0) * 1.5

    return doc

# --- MAIN STREAMLIT APP ---
st.set_page_config(page_title="SLD Automator", layout="wide")
st.title("⚡ Power System SLD Generator")

uploaded_file = st.file_uploader("Upload Substation Excel File", type=["xlsx"])

if uploaded_file is not None:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(uploaded_file.getvalue())
        file_path = tmp.name

    try:
        df = pd.read_excel(file_path, header=None)
        num_feeders = int(df.iloc[0, 1])
        feeder_types = df.iloc[1, 2 : 2 + num_feeders].tolist()

        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active 
        b6 = sheet["B6"].value
        b8 = sheet["B8"].value
        b9 = sheet["B9"].value
        d6 = sheet["D6"].value
        d7 = sheet["D7"].value

        voltage_value = sheet["B12"].value
        b12 = int(voltage_value) if voltage_value else 0

        if b6 in ["Double Main Transfer Bus", "Double Main Bus"]:
            feeder_types = [
                "Line_Bay" if pd.isna(f) or str(f).strip() == "" or str(f).strip() == "Future_Bay"
                else str(f).strip()
                for f in feeder_types
            ]
        else:
            feeder_types = [
                "Line_Bay" if pd.isna(f) or str(f).strip() == "" or str(f).strip() == "Bus_Coupler" or str(f).strip() == "Cable Feeder" or str(f).strip() == "Transfer_Bus_coupler"
                else str(f).strip()
                for f in feeder_types
            ]

        feeder_names = list(df.iloc[2, 2:2+num_feeders]) 

        def get_fontsize(num_feeders):
            if b6 == "Double Main Transfer Bus" :
                if num_feeders <= 3: return 3
                else: return 3
            else:
                return 4

        def draw_breaker(ax, x, y, label, fs):
            ax.add_patch(Rectangle((x-0.1, y-0.2), 0.2, 0.4, fill=False, edgecolor='black', linewidth=0.5))
            ax.plot([x, x], [y+.2, y+1], color='red', linewidth=0.5)
            ax.plot([x, x], [y-.2, y-1], color='red', linewidth=0.5)
            ax.text(x+.3, y, label, fontsize=fs, ha='center')

        def draw_breaker_coupler(ax, x, y, label, fs):
            ax.add_patch(Rectangle((x-0.1, y-0.2), 0.2, 0.4, fill=False, edgecolor='black', linewidth=0.5))
            ax.plot([x-.45, x-.1], [y, y], color='red', linewidth=0.5)
            ax.plot([x+.1, x+.45], [y, y], color='red', linewidth=0.5)
            ax.text(x, y+.5, label, fontsize=fs, ha='center')

        def draw_isolator(ax, x, y, label, fs):
            if b6 == "Double Main Transfer Bus" :
                ax.plot([x, x], [y-.12, y+.2], color='red', linewidth=0.5)
                ax.text(x+.2, y-.2, label, fontsize=fs, ha='center')
            else:
                ax.plot([x, x], [y-.12, y+.4], color='red', linewidth=0.5)
                ax.text(x+.3, y-.3, label, fontsize=fs, ha='center')

            arc1 = Arc((x , y-0.6/4), width=0.04, height=0.08, angle=0, theta1=0, theta2=360, color='red', linewidth=0.5)
            ax.add_patch(arc1)
            arc2 = Arc((x , y-.5-0.6/4), width=0.04, height=0.08, angle=0, theta1=0, theta2=360, color='red', linewidth=0.5)
            ax.add_patch(arc2)

            ax.plot([x-.05, x+.05], [y-.55, y-.25], color='red', linewidth=0.5)
            ax.plot([x, x], [y-.7, y-1.2], color='red', linewidth=0.5)

        def earth_sh(ax, x, y, label, fs):
            y = y - .2
            ax.plot([x, x-.2], [y-.6, y-.6], color='red', linewidth=0.5)
            arc1 = Arc((x-.2 , y-.45-0.6/4), width=0.04, height=0.08, angle=0, theta1=0, theta2=360, color='red', linewidth=0.5)
            ax.add_patch(arc1)
            arc1 = Arc((x-.35 , y-.45-0.6/4), width=0.04, height=0.08, angle=0, theta1=0, theta2=360, color='green', linewidth=0.5)
            ax.add_patch(arc1)
            ax.plot([x-.2, x-.35], [y-.35, y-.6], color='green', linewidth=0.5)
            ax.plot([x-.35, x-.5], [y-.6, y-.6], color='green', linewidth=0.5)
            ax.plot([x-.5, x-.5], [y-.45, y-.75], color='green', linewidth=0.5)
            ax.plot([x-.55, x-.55], [y-.5, y-.7], color='green', linewidth=0.5)
            ax.plot([x-.6, x-.6], [y-.55, y-.65], color='green', linewidth=0.5)
            ax.text(x-.4, y-.35, label, fontsize=fs, ha='center')

        def draw_ct(ax, x, y, label, fs):
            spacing = 0.75
            arc1 = Arc((x+.03 , y- spacing/4), width=0.2, height=0.4, angle=0, theta1=80, theta2=280, color='blue', linewidth=0.5)
            arc2 = Arc((x+.03 , y+ spacing/4), width=0.2, height=0.4, angle=0, theta1=80, theta2=280, color='blue', linewidth=0.5)
            ax.add_patch(arc1)
            ax.add_patch(arc2)
            ax.text(x-.3, y, label, fontsize=fs, ha='center')

        def draw_wt(ax, x, y, label, fs):
            arc1 = Arc((x , y-0.6/4), width=0.2, height=0.4, angle=0, theta1=90, theta2=360, color='red', linewidth=0.5)
            ax.add_patch(arc1)
            ax.plot([x, x+.1], [y-.15, y-.15], color='red', linewidth=0.5)
            ax.text(x-.3, y, label, fontsize=fs, ha='center')

        def draw_cvt(ax, x, y, label, fs):
            y = y - .2
            ax.plot([x, x + .15], [y - .6, y - .6], color='red', linewidth=0.5)
            ax.plot([x + .15, x + .15], [y - .45, y - .75], color='red', linewidth=0.5)
            ax.plot([x + .2, x + .2], [y - .45, y - .75], color='red', linewidth=0.5)
            ax.plot([x + .2, x + .55], [y - .6, y - .6], color='red', linewidth=0.5)
            ax.plot([x + .55, x + .55], [y - .45, y - .75], color='red', linewidth=0.5)
            ax.plot([x + .6, x + .6], [y - .45, y - .75], color='red', linewidth=0.5)
            ax.plot([x + .6, x + .7], [y - .6, y - .6], color='red', linewidth=0.5)
            ax.plot([x + .7, x + .7], [y - .45, y - .75], color='green', linewidth=0.5)
            ax.plot([x + .75, x + .75], [y - .5, y - .7], color='green', linewidth=0.5)
            ax.plot([x + .8, x + .8], [y - .55, y - .65], color='green', linewidth=0.5)
            ax.plot([x + .375, x + .375], [y - .6, y - 1.4], color='red', linewidth=0.5)
            ax.plot([x + .375, x + .45], [y - 1.4, y - 1.4], color='red', linewidth=0.5)
            arc1 = Arc((x+.45 , y-1.4-0.45/4), width=0.1, height=0.2, angle=0, theta1=270, theta2=90, color='red', linewidth=0.5)
            arc2 = Arc((x+.45 , y-1.4+0.45/4), width=0.1, height=0.2, angle=0, theta1=270, theta2=90, color='red', linewidth=0.5)
            ax.add_patch(arc1)
            ax.add_patch(arc2)
            ax.plot([x + .575, x + .575], [y - 1, y - 1.7], color='red', linewidth=.5)
            ax.plot([x + .55, x + .55], [y - 1, y - 1.7], color='red', linewidth=.5)

            spacing = 0.25  
            arc1 = Arc((x+.675 , y-1.1- spacing/4), width=0.1, height=0.2, angle=0, theta1=80, theta2=280, color='red', linewidth=.5)
            ax.add_patch(arc1)
            arc2 = Arc((x+.675 , y-1.025+ spacing/4), width=0.1, height=0.2, angle=0, theta1=80, theta2=280, color='red', linewidth=.5)
            ax.add_patch(arc2)
            arc3 = Arc((x+.675 , y-1.6+ spacing/4), width=0.1, height=0.2, angle=0, theta1=80, theta2=280, color='red', linewidth=.5)
            ax.add_patch(arc3)
            ax.text(x + .4, y - .4, label, fontsize=fs, ha='center')
            arc4 = Arc((x+.675 , y-1.8+ spacing/4), width=0.1, height=0.2, angle=0, theta1=80, theta2=280, color='red', linewidth=.5)
            ax.add_patch(arc4)

        def draw_symbol(ax, x, y, label, fs):
            triangle = Polygon([[x, y-0.1], [x+0.1, y+0.1], [x-0.1, y+0.1]], closed=True, fill=False, edgecolor='red', linewidth=0.5)
            ax.add_patch(triangle)

        def draw_symbol_upp(ax, x, y, label, fs):
            triangle = Polygon([[x, y+0.1], [x+0.1, y-0.1], [x-0.1, y-0.1]], closed=True, fill=False, edgecolor='red', linewidth=0.5)
            ax.add_patch(triangle)

        def draw_name(ax, x, y, label, fs):
            ax.text(x, y+.5, label, fontsize=fs+2, ha='center')

        def draw_la(ax, x, y, label, fs):
            y = y - .2
            ax.plot([x-.9, x], [y, y], color='red', linewidth=0.5)
            ax.plot([x - .9, x - .9], [y +.2, y -.2], color='green', linewidth=0.5)
            ax.plot([x - .95, x - .95], [y +.15, y -.15], color='green', linewidth=0.5)
            ax.plot([x - 1, x - 1], [y +.1, y -.1], color='green', linewidth=0.5)
            ax.text(x-.5, y-0.5, label, fontsize=fs, ha='center')

        def la_comp(ax, x, y):
            ax.add_patch(Polygon([[x+0.1, y+0.1], [x-0.1, y], [x+0.1, y-0.1]], closed=True, fill=True,color='red', linewidth=0.5))
            ax.add_patch(Rectangle((x-0.2, y-0.2), 0.4, 0.4, fill=False, edgecolor='red', linewidth=0.5))

        def draw_ict(ax, x, y, label, fs):
            arc1 = Arc((x , y-0.6/4), width=0.3, height=0.6, angle=0, theta1=0, theta2=360, color='red', linewidth=0.5)
            ax.add_patch(arc1)
            arc2 = Arc((x-.15 , y-0.6/4), width=0.2, height=0.4, angle=0, theta1=0, theta2=360, color='red', linewidth=0.5)
            ax.add_patch(arc2)
            arc3 = Arc((x , (y-.255-0.6/4)), width=0.4, height=1.2, angle=0, theta1=270, theta2=90, color='red', linewidth=.5)
            ax.add_patch(arc3)
            ax.text(x-.25, y-.7, label, fontsize=fs, ha='center')

        def draw_ict_upp(ax, x, y, label, fs):
            arc1 = Arc((x , y-.525-0.6/4), width=0.3, height=0.6, angle=0, theta1=0, theta2=360, color='red', linewidth=0.5)
            ax.add_patch(arc1)
            arc2 = Arc((x-.15 , y-.525-0.6/4), width=0.2, height=0.4, angle=0, theta1=0, theta2=360, color='red', linewidth=0.5)
            ax.add_patch(arc2)
            arc3 = Arc((x , (y-.255-0.6/4)), width=0.4, height=1.2, angle=0, theta1=270, theta2=90, color='red', linewidth=.5)
            ax.add_patch(arc3)
            ax.text(x-.3, y-.2, label, fontsize=fs, ha='center')

        def draw_reacter(ax, x, y, label, fs):
            arc1 = Arc((x-.025 , y- .2/4), width=0.2, height=0.4, angle=0, theta1=80, theta2=300, color='red', linewidth=0.5)
            arc2 = Arc((x-.025 , y-1.2/4), width=0.2, height=0.4, angle=0, theta1=60, theta2=300, color='red', linewidth=0.5)
            ax.add_patch(arc1)
            ax.add_patch(arc2)
            arc3 = Arc((x-.025 , y-2.2/4), width=0.2, height=0.4, angle=0, theta1=60, theta2=300, color='red', linewidth=0.5)
            ax.add_patch(arc3)
            arc4 = Arc((x-.025 , y-3.2/4), width=0.2, height=0.4, angle=0, theta1=60, theta2=280, color='red', linewidth=0.5)
            ax.add_patch(arc4)
            ax.text(x-.45, y-.8, label, fontsize=fs, ha='center')

        def draw_earth_symbol(ax, x, y, label, fs):
            ax.plot([x-.125, x+.125], [y+.1, y+.1], color='green', linewidth=0.5)
            ax.plot([x-.1, x+.1], [y, y], color='green', linewidth=0.5)
            ax.plot([x-.075, x+.075], [y-.1, y-.1], color='green', linewidth=0.5)
            ax.plot([x-.05, x+.05], [y-.2, y-.2], color='green', linewidth=0.5)
            ax.plot([x-.025, x+.025], [y-.3, y-.3], color='green', linewidth=0.5)

        def draw_earth_symbol_upp(ax, x, y, label, fs):
            ax.plot([x-.125, x+.125], [y-.3, y-.3], color='green', linewidth=0.5)
            ax.plot([x-.1, x+.1], [y-.2, y-.2], color='green', linewidth=0.5)
            ax.plot([x-.075, x+.075], [y-.1, y-.1], color='green', linewidth=0.5)
            ax.plot([x-.05, x+.05], [y, y], color='green', linewidth=0.5)
            ax.plot([x-.025, x+.025], [y+.1, y+.1], color='green', linewidth=0.5)

        # -------------------------------------------------------------
        # BRANCH 1: Double Main / Transfer Bus
        # -------------------------------------------------------------
        if b6 in ["Double Main Transfer Bus", "Double Main Bus"]:

            def get_labels(feeder_num):
                voltage_value = sheet["B12"].value
                b12 = int(voltage_value) if voltage_value else 0
                b13 = sheet["B13"].value
                b14 = sheet["B14"].value
                b15 = sheet["B15"].value
                b16 = sheet["B16"].value
                b17 = sheet["B17"].value
                b18 = sheet["B18"].value
                b19 = sheet["B19"].value
                b20 = sheet["B20"].value
                b21 = sheet["B21"].value
                b22 = sheet["B22"].value
                b23 = sheet["B23"].value
                b24 = sheet["B24"].value
                b25 = sheet["B25"].value
                b26 = sheet["B26"].value
                b27 = sheet["B27"].value
                d14 = sheet["D14"].value
                d15 = sheet["D15"].value
                d16 = sheet["D16"].value
                d17 = sheet["D17"].value
                d18 = sheet["D18"].value
                
                if b13 == "Predefined":
                    labels = {
                        "base_isolator" : f"{b12+feeder_num}89A",
                        "base_isolatorb" : f"{b12+feeder_num}89B",
                        "breaker_lbl" : f"{b12+feeder_num}52",
                        "earth_lbl1" : f"{b12+feeder_num}89AE",
                        "earth_lbl2" : f"{b12+feeder_num}89CE1",
                        "iso_lbl2" : f"{b12+feeder_num}89C",
                        "iso_lbl3" : f"{b12+feeder_num}89D",
                        "ct_lbl" : f"{int((b12+feeder_num))}CT",
                        "wt_lbl" : f"{int((b12+feeder_num))}WT",
                        "cvt_lbl" : f"{int((b12+feeder_num))}CVT",
                        "la_lbl" : f"{int((b12+feeder_num))}LA",
                        "symbol_lbl" : f"{int((b12+feeder_num))}",
                        "earth_lbl3" : f"{b12+feeder_num}89CE2",
                        "ict_lbl" : f"{int((b12+feeder_num))}ICT",
                        "rect_lbl" : f"{int((b12+feeder_num))}Reactor",
                        "symbol_lbl_Tcup" : f"{int((b12+feeder_num))} \n (TFR_Bus_coupler)",
                        "earth_lbl1_Bcup" : f"{b12+feeder_num}89AE1",
                        "earth_lbl2_Bcup" : f"{b12+feeder_num}89AE2",
                        "earth_lbl3_Bcup" : f"{b12+feeder_num}89BE1",
                        "earth_lbl4_Bcup" : f"{b12+feeder_num}89BE2",
                        "symbol_lbl_Bcup" : f"{int((b12+feeder_num))} \n (Bus_coupler)",
                        "No_bay_lbl" : f"{int((b12+feeder_num))} \n (Not Exist)"
                    }
                    return labels
                else:
                    labels = {
                        "base_isolator" : b14,
                        "base_isolatorb" : b15,
                        "breaker_lbl" : b17,
                        "earth_lbl1" : b16,
                        "earth_lbl2" : b18,
                        "iso_lbl2" : b19,
                        "iso_lbl3" : b21,
                        "ct_lbl" : b22,
                        "wt_lbl" : b23,
                        "cvt_lbl" : b24,
                        "la_lbl" : b25,
                        "symbol_lbl" : f"{int((b12+feeder_num))}",
                        "earth_lbl3" : b20,
                        "ict_lbl" : b26,
                        "rect_lbl" : b27,
                        "symbol_lbl_Tcup" : f"{int((b12+feeder_num))} \n (TFR_Bus_coupler)",
                        "earth_lbl1_Bcup" : d14,
                        "earth_lbl2_Bcup" : d15,
                        "earth_lbl3_Bcup" : d16,
                        "earth_lbl4_Bcup" : d17,
                        "symbol_lbl_Bcup" : f"{int((b12+feeder_num))} \n (Bus_coupler)",
                        "No_bay_lbl" : d18
                    }
                    return labels

            def draw_feeder1(ax, x_offset, y_offset, feeder_num, fs):
                L = get_labels(feeder_num)
                ax.plot([x_offset, x_offset], [y_offset, y_offset+1.8], color='red', linewidth=0.5)
                ax.plot([x_offset+0.5, x_offset+0.5], [y_offset, y_offset+0.8], color='red', linewidth=0.5)
                ax.plot([x_offset, x_offset+0.5], [y_offset-1.2, y_offset-1.2], color='red', linewidth=0.5)
                draw_isolator(ax, x_offset, y_offset, L["base_isolator"], fs)
                draw_isolator(ax, x_offset+0.5, y_offset, L["base_isolatorb"], fs)
                earth_sh(ax, x_offset, y_offset, L["earth_lbl1"], fs)
                draw_breaker(ax, x_offset+0.25, y_offset-2.2, L["breaker_lbl"], fs)

                if b6=="Double Main Transfer Bus":
                    ax.plot([x_offset+.25, x_offset+0.5], [y_offset-4.5, y_offset-4.5], color='red', linewidth=0.5)
                    ax.plot([x_offset+0.5, x_offset+0.5], [y_offset-4.5, y_offset-7.5], color='red', linewidth=0.5)
                    ax.plot([x_offset+0.5, x_offset+0.5], [y_offset-7.75, y_offset-10], color='red', linewidth=0.5)
                    earth_sh(ax, x_offset+0.25, y_offset-2.2, L["earth_lbl2"], fs)
                    draw_isolator(ax, x_offset+0.25, y_offset-3.2, L["iso_lbl2"], fs)
                    earth_sh(ax, x_offset+0.25, y_offset-3.2, L["earth_lbl3"], fs)
                    draw_isolator(ax, x_offset+0.25, y_offset-4.6, L["iso_lbl3"], fs)
                    draw_ct(ax, x_offset+0.5, y_offset-6.7, L["ct_lbl"], fs)
                    if f_type=="Cable Feeder":
                        ax.plot([x_offset+0.5, x_offset+0.5], [y_offset-7.5, y_offset-7.75], color='red', linewidth=0.5)
                    else:
                        draw_wt(ax, x_offset+0.5, y_offset-7.6, L["wt_lbl"], fs)
                    draw_cvt(ax, x_offset+0.5, y_offset-7.7, L["cvt_lbl"], fs)
                    draw_la(ax, x_offset+0.5, y_offset-9, L["la_lbl"], fs)
                    la_comp(ax, x_offset, y_offset-9.2)
                    draw_name(ax, x_offset+0.25, y_offset+1.75,L["symbol_lbl"], fs)
                    draw_symbol(ax, x_offset+0.5, y_offset-10.1, L["symbol_lbl"], fs)

                elif b6=="Double Main Bus":
                    ax.plot([x_offset+.25, x_offset+0.25], [y_offset-2.7, y_offset-4.5], color='red', linewidth=0.5)
                    draw_ct(ax, x_offset+0.25, y_offset-3.4, L["ct_lbl"], fs)
                    earth_sh(ax, x_offset+0.25, y_offset-3.7, L["earth_lbl2"], fs)
                    draw_isolator(ax, x_offset+0.25, y_offset-4.7, L["iso_lbl2"], fs)
                    earth_sh(ax, x_offset+0.25, y_offset-4.85, L["earth_lbl3"], fs)
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset-5.9, y_offset-6.55], color='red', linewidth=0.5)
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset-6.75, y_offset-10.4], color='red', linewidth=0.5)
                    if f_type=="Cable Feeder":
                        ax.plot([x_offset+0.25, x_offset+0.25], [y_offset-6.55, y_offset-6.75], color='red', linewidth=0.5)
                    else:
                        draw_wt(ax, x_offset+0.25, y_offset-6.6, L["wt_lbl"], fs)
                    draw_cvt(ax, x_offset+0.25, y_offset-7, L["cvt_lbl"], fs)
                    draw_la(ax, x_offset+0.25, y_offset-8.5, L["la_lbl"], fs)
                    la_comp(ax, x_offset-0.25, y_offset-8.7)
                    draw_name(ax, x_offset+0.25, y_offset+1.75,L["symbol_lbl"], fs)
                    draw_symbol(ax, x_offset+0.25, y_offset-10.5, L["symbol_lbl"], fs)

            def draw_feeder2(ax, x_offset, y_offset, feeder_num, fs):
                L = get_labels(feeder_num)
                ax.plot([x_offset, x_offset], [y_offset, y_offset+1.8], color='red', linewidth=0.5)
                ax.plot([x_offset+0.5, x_offset+0.5], [y_offset, y_offset+0.8], color='red', linewidth=0.5)
                ax.plot([x_offset, x_offset+0.5], [y_offset-1.2, y_offset-1.2], color='red', linewidth=0.5)
                draw_isolator(ax, x_offset, y_offset, L["base_isolator"], fs)
                draw_isolator(ax, x_offset+0.5, y_offset, L["base_isolatorb"], fs)
                earth_sh(ax, x_offset, y_offset, L["earth_lbl1"], fs)
                draw_breaker(ax, x_offset+0.25, y_offset-2.2, L["breaker_lbl"], fs)

                if b6=="Double Main Transfer Bus":
                    ax.plot([x_offset+.25, x_offset+0.5], [y_offset-4.5, y_offset-4.5], color='red', linewidth=0.5)
                    ax.plot([x_offset+0.5, x_offset+0.5], [y_offset-4.5, y_offset-8.85], color='red', linewidth=0.5)
                    ax.plot([x_offset+0.5, x_offset+0.5], [y_offset-10, y_offset-11.1], color='red', linewidth=0.5)
                    earth_sh(ax, x_offset+0.25, y_offset-2.2, L["earth_lbl2"], fs)
                    draw_isolator(ax, x_offset+0.25, y_offset-3.2, L["iso_lbl2"], fs)
                    earth_sh(ax, x_offset+0.25, y_offset-3.2, L["earth_lbl3"], fs)
                    draw_isolator(ax, x_offset+0.25, y_offset-4.6, L["iso_lbl3"], fs)
                    draw_ct(ax, x_offset+0.5, y_offset-6.7, L["ct_lbl"], fs)
                    draw_la(ax,x_offset+0.5, y_offset-7.6, L["la_lbl"], fs)
                    la_comp(ax, x_offset, y_offset-7.8)
                    draw_ict(ax,x_offset+0.5, y_offset-9, L["ict_lbl"], fs)
                    draw_symbol(ax, x_offset+0.5, y_offset-11.2, L["symbol_lbl"], fs)
                    draw_name(ax, x_offset+0.25, y_offset+1.75, L["symbol_lbl"], fs)

                elif b6=="Double Main Bus":
                    ax.plot([x_offset+.25, x_offset+0.25], [y_offset-2.7, y_offset-4.5], color='red', linewidth=0.5)
                    draw_ct(ax, x_offset+0.25, y_offset-3.4, L["ct_lbl"], fs)
                    earth_sh(ax, x_offset+0.25, y_offset-3.7, L["earth_lbl2"], fs)
                    draw_isolator(ax, x_offset+0.25, y_offset-4.7, L["iso_lbl2"], fs)
                    earth_sh(ax, x_offset+0.25, y_offset-4.85, L["earth_lbl3"], fs)
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset-5.9, y_offset-7.8], color='red', linewidth=0.5)
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset-9, y_offset-10.1], color='red', linewidth=0.5)
                    draw_la(ax,x_offset+0.25, y_offset-6.6, L["la_lbl"], fs)
                    la_comp(ax, x_offset-.25, y_offset-6.8)
                    draw_ict(ax,x_offset+0.25, y_offset-8, L["ict_lbl"], fs)
                    draw_symbol(ax, x_offset+0.25, y_offset-10.2, L["symbol_lbl"], fs)
                    draw_name(ax, x_offset+0.25, y_offset+1.75, L["symbol_lbl"], fs)

            def draw_feeder3(ax, x_offset, y_offset, feeder_num, fs):
                L = get_labels(feeder_num)
                ax.plot([x_offset, x_offset], [y_offset, y_offset+1.8], color='red', linewidth=0.5)
                ax.plot([x_offset+0.5, x_offset+0.5], [y_offset, y_offset+0.8], color='red', linewidth=0.5)
                ax.plot([x_offset, x_offset+0.5], [y_offset-1.2, y_offset-1.2], color='red', linewidth=0.5)
                draw_isolator(ax, x_offset, y_offset, L["base_isolator"], fs)
                draw_isolator(ax, x_offset+0.5, y_offset, L["base_isolatorb"], fs)
                earth_sh(ax, x_offset, y_offset, L["earth_lbl1"], fs)
                draw_breaker(ax, x_offset+0.25, y_offset-2.2, L["breaker_lbl"], fs)

                if b6=="Double Main Transfer Bus":
                    ax.plot([x_offset+.25, x_offset+0.5], [y_offset-4.5, y_offset-4.5], color='red', linewidth=0.5)
                    ax.plot([x_offset+0.5, x_offset+0.5], [y_offset-4.5, y_offset-8.85], color='red', linewidth=0.5)
                    ax.plot([x_offset+0.5, x_offset+0.5], [y_offset-10, y_offset-11.1], color='red', linewidth=0.5)
                    earth_sh(ax, x_offset+0.25, y_offset-2.2, L["earth_lbl2"], fs)
                    draw_isolator(ax, x_offset+0.25, y_offset-3.2, L["iso_lbl2"], fs)
                    earth_sh(ax, x_offset+0.25, y_offset-3.2, L["earth_lbl3"], fs)
                    draw_isolator(ax, x_offset+0.25, y_offset-4.6, L["iso_lbl3"], fs)
                    draw_ct(ax, x_offset+0.5, y_offset-6.7, L["ct_lbl"], fs)
                    draw_la(ax,x_offset+0.5, y_offset-7.6, L["la_lbl"], fs)
                    la_comp(ax, x_offset, y_offset-7.8)
                    draw_reacter(ax,x_offset+0.5, y_offset-9, L["rect_lbl"], fs)
                    draw_earth_symbol(ax, x_offset+0.5, y_offset-11.2, L["symbol_lbl"], fs)
                    draw_name(ax, x_offset+0.25, y_offset+1.75, L["symbol_lbl"], fs)

                elif b6=="Double Main Bus":
                    ax.plot([x_offset+.25, x_offset+0.25], [y_offset-2.7, y_offset-4.5], color='red', linewidth=0.5)
                    draw_ct(ax, x_offset+0.25, y_offset-3.4, L["ct_lbl"], fs)
                    earth_sh(ax, x_offset+0.25, y_offset-3.7, L["earth_lbl2"], fs)
                    draw_isolator(ax, x_offset+0.25, y_offset-4.7, L["iso_lbl2"], fs)
                    earth_sh(ax, x_offset+0.25, y_offset-4.85, L["earth_lbl3"], fs)
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset-5.9, y_offset-7.85], color='red', linewidth=0.5)
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset-9, y_offset-10.1], color='red', linewidth=0.5)
                    draw_la(ax,x_offset+0.25, y_offset-6.6, L["la_lbl"], fs)
                    la_comp(ax, x_offset-.25, y_offset-6.8)
                    draw_reacter(ax,x_offset+0.25, y_offset-8, L["rect_lbl"], fs)
                    draw_earth_symbol(ax, x_offset+0.25, y_offset-10.2, L["symbol_lbl"], fs)
                    draw_name(ax, x_offset+0.25, y_offset+1.75, L["symbol_lbl"], fs)

            def draw_feeder4(ax, x_offset, y_offset, feeder_num, fs):
                L = get_labels(feeder_num)
                if b6=="Double Main Transfer Bus":
                    ax.plot([x_offset, x_offset], [y_offset, y_offset+1.8], color='red', linewidth=0.5)
                    ax.plot([x_offset+.5, x_offset+.5], [y_offset, y_offset+0.8], color='red', linewidth=0.5)
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset-3.2, y_offset-4.6], color='red', linewidth=0.5)
                    draw_isolator(ax, x_offset, y_offset,L["base_isolator"], fs)
                    draw_isolator(ax, x_offset+0.5, y_offset, L["base_isolatorb"], fs)
                    earth_sh(ax, x_offset, y_offset, L["earth_lbl1"], fs)
                    ax.plot([x_offset, x_offset+0.5], [y_offset-1.2, y_offset-1.2], color='red', linewidth=0.5)
                    draw_breaker(ax, x_offset+0.25, y_offset-2.2, L["breaker_lbl"], fs)
                    earth_sh(ax, x_offset+0.25, y_offset-3.6, L["earth_lbl2"], fs)
                    draw_isolator(ax, x_offset+0.25, y_offset-4.6, L["iso_lbl2"], fs)
                    earth_sh(ax, x_offset+0.25, y_offset-4.6, L["earth_lbl3"], fs)
                    draw_ct(ax, x_offset+0.25, y_offset-3.5, L["ct_lbl"], fs)
                    draw_name(ax, x_offset+0.25, y_offset+1.75, L["symbol_lbl_Tcup"], fs)
                elif b6=="Double Main Bus":
                    draw_feeder1(ax, x_pos, y_start, feeder_label, fontsize)

            def draw_feeder5(ax, x_offset, y_offset, feeder_num, fs):
                L = get_labels(feeder_num)
                ax.plot([x_offset-.2, x_offset-.2], [y_offset-.3, y_offset+1.8], color='red', linewidth=0.5)
                ax.plot([x_offset+0.7, x_offset+.7], [y_offset-.3, y_offset+0.8], color='red', linewidth=0.5)
                ax.plot([x_offset-.2, x_offset-.2], [y_offset-4, y_offset-1.5], color='red', linewidth=0.5)
                ax.plot([x_offset+0.7, x_offset+.7], [y_offset-4, y_offset-1.5], color='red', linewidth=0.5)
                draw_isolator(ax, x_offset-.2, y_offset-.3, L["base_isolator"], fs)
                draw_isolator(ax, x_offset+0.7, y_offset-.3, L["base_isolatorb"], fs)
                earth_sh(ax, x_offset-0.2, y_offset+.8, L["earth_lbl1_Bcup"], fs)
                earth_sh(ax, x_offset-0.2, y_offset-.5, L["earth_lbl2_Bcup"], fs)
                earth_sh(ax, x_offset+.7, y_offset+.8, L["earth_lbl3_Bcup"], fs)
                earth_sh(ax, x_offset+.7, y_offset-.5, L["earth_lbl4_Bcup"], fs)
                draw_ct(ax, x_offset-0.2, y_offset-2.5, L["ct_lbl"], fs)
                draw_breaker_coupler(ax, x_offset+0.25, y_offset-4, L["breaker_lbl"], fs)
                draw_name(ax, x_offset+0.25, y_offset+1.75, L["symbol_lbl_Bcup"], fs)

            def draw_feeder6(ax, x_offset, y_offset, feeder_num, fs):
                L = get_labels(feeder_num)
                draw_name(ax, x_offset+0.25, y_offset+1.75, L["No_bay_lbl"], fs)

            fig_width = max(12, num_feeders * 2)
            fig, ax = plt.subplots(figsize=(fig_width, 6))
            x_start = 5
            y_start = 8.2
            gap = 2
            fontsize = get_fontsize(num_feeders)

            for i in range(num_feeders):
                f_type = feeder_types[i] if i < len(feeder_types) else '1'
                f_name = str(feeder_names[i]) if i < len(feeder_names) else ''
                if f_name.lower() == 'nan':  
                    f_name = ''
                x_pos = x_start + i * gap
                feeder_label = i + 1

                ax.plot([1, num_feeders*3], [10, 10], color='blue', linewidth=0.5)
                ax.text(1.5, 10.5, f"{b12} KV_BUS1", fontsize=fontsize+4,  va='center')
                ax.plot([1, num_feeders*3], [9, 9], color='green', linewidth=0.5)
                ax.text(1.5, 9.5, f"{b12} KV_BUS2", fontsize=fontsize+4,  va='center')
                if b6=="Double Main Transfer Bus":
                    ax.plot([1, num_feeders*3], [2.4, 2.4], color='black', linewidth=0.5)
                    ax.text(1.5, 2.4+.5, f"{b12} KV_Transfer BUS", fontsize=fontsize+4,  va='center')

                if f_type == 'Line_Bay': draw_feeder1(ax, x_pos, y_start, feeder_label, fontsize)
                elif f_type == 'ICT' : draw_feeder2(ax, x_pos, y_start, feeder_label, fontsize)
                elif f_type == 'Reactor' : draw_feeder3(ax, x_pos, y_start, feeder_label, fontsize)
                elif f_type == 'Transfer_Bus_coupler' : draw_feeder4(ax, x_pos, y_start, feeder_label, fontsize)
                elif f_type == 'Bus_Coupler' : draw_feeder5(ax, x_pos, y_start, feeder_label, fontsize)
                elif f_type == 'No_bay' : draw_feeder6(ax, x_pos, y_start, feeder_label, fontsize)
                elif f_type == 'Cable Feeder' : draw_feeder1(ax, x_pos, y_start, feeder_label, fontsize)
                else: draw_feeder1(ax, x_pos, y_start, feeder_label, fontsize)

                words = f_name.split()
                lines, current_line = [], ""
                for word in words:
                    if len(current_line + " " + word) <= 24:
                        if current_line: current_line += " " + word
                        else: current_line = word
                    else:
                        lines.append(current_line)
                        current_line = word
                if current_line: lines.append(current_line)
                wrapped_text = "\n".join(lines)
                label_text = f"{b12 + feeder_label}"
                full_text = f"{label_text}\n{wrapped_text}"

                if b6=="Double Main Transfer Bus":
                    ax.text(x_pos+.5, y_start-12, full_text, ha='center', va='top', fontsize=fontsize+1)
                elif b6=="Double Main Bus":
                    ax.text(x_pos+.25, y_start-11.25, full_text, ha='center', va='top', fontsize=fontsize+1)

            center_x = (x_start + (num_feeders-1)*gap)/2
            ax.text(center_x, 14, 'POWERGRID CORPORATION OF INDIA LTD', fontsize=fontsize+25, va='center', ha='center')  
            ax.text(center_x, 12, b9, fontsize=fontsize+15, va='center', ha='center')  

            ax.set_xlim(1,(num_feeders)*2+6)
            ax.set_ylim(-6*1.05,1.5*10.5)
            ax.axis('off')

            img_path = os.path.join(os.path.dirname(file_path), "temp_sld.png")
            fig.savefig(img_path, bbox_inches='tight', dpi=300)

            if "Feeder Diagram" in wb.sheetnames:
                std = wb["Feeder Diagram"]
                wb.remove(std)
            ws_new = wb.create_sheet("Feeder Diagram")
            img = XLImage(img_path)
            img.width = num_feeders*100*2
            img.height = 250*1.8*2
            ws_new.add_image(img, "A1")
            
            wb.save(file_path)
            wb.close()

            # Generate Outputs
            st.pyplot(fig)
            pdf_io = io.BytesIO()
            fig.savefig(pdf_io, format='pdf', bbox_inches='tight')
            dxf_doc = export_ax_to_dxf(ax)
            dxf_io = io.StringIO()
            dxf_doc.write(dxf_io)

            st.success("Generation Complete!")
            col1, col2, col3 = st.columns(3)
            with col1:
                with open(file_path, "rb") as f:
                    st.download_button(label="📥 Download Excel", data=f, file_name="SLD_Integrated.xlsx")
            with col2:
                st.download_button(label="📥 Download PDF", data=pdf_io.getvalue(), file_name="Substation_SLD.pdf", mime="application/pdf")
            with col3:
                st.download_button(label="📥 Download AutoCAD DXF", data=dxf_io.getvalue(), file_name="Substation_SLD.dxf", mime="application/dxf")
            
            os.remove(img_path)

        # -------------------------------------------------------------
        # BRANCH 2: One and a half Busbar (Else condition)
        # -------------------------------------------------------------
        else:
            def get_labels(feeder_num):
                voltage_value = sheet["B12"].value
                b12=int(voltage_value) if voltage_value else 0
                b13 = sheet["B13"].value
                b14 = sheet["B14"].value
                b15 = sheet["B15"].value
                b16 = sheet["B16"].value
                b17 = sheet["B17"].value
                b18 = sheet["B18"].value
                b19 = sheet["B19"].value
                b20 = sheet["B20"].value
                b21 = sheet["B21"].value
                b22 = sheet["B22"].value
                b23 = sheet["B23"].value
                b24 = sheet["B24"].value
                b25 = sheet["B25"].value
                b26 = sheet["B26"].value
                b27 = sheet["B27"].value
                e6 = sheet["E6"].value
                e7 = sheet["E7"].value

                current_type = feeder_types[i] if i < len(feeder_types) else ""
                bay_num = i + 1 

                if b13 == "Predefined":
                    if (i + 1 == d6) or (i + 1 == d7):
                        earth_label = f"{b12+feeder_num}89AE2"
                    else:
                        earth_label = f"{b12+feeder_num}89AE"

                    if (i + 1 == d6) or (i + 1 == d7):
                        earth_label2 = f"{b12+feeder_num}89AE1"
                    else:
                        earth_label2 = f"{b12+feeder_num}89AE1"

                    if current_type in ["Line_Bay", "Future_Bay"] and ((i + 1) % 3 == 1 or (i + 1) % 3 == 0):
                        pair_index = i - 2 if (i + 1) % 3 == 0 else i + 2  
                        pair_index = pair_index if pair_index < len(feeder_types) else None

                        if pair_index is not None:
                            pair_type = feeder_types[pair_index]
                            if pair_type in ["Line_Bay", "Future_Bay"]:
                                if (bay_num - 1) % 3 == 0:     
                                    ct_label = f"{b12+feeder_num+1}BCT"
                                elif bay_num % 3 == 0:          
                                    ct_label = f"{b12+feeder_num-1}ACT"
                                else:
                                    ct_label = f"{b12+feeder_num}CT"
                            else:
                                if (bay_num - 1) % 3 == 0:     
                                    ct_label = f"{b12+feeder_num+1}CT"
                                elif bay_num % 3 == 0:          
                                    ct_label = f"{b12+feeder_num-1}CT"
                                else:
                                    ct_label = f"{b12+feeder_num}CT"
                        else:
                            if (bay_num - 1) % 3 == 0:
                                ct_label = f"{b12+feeder_num+1}CT"
                            elif bay_num % 3 == 0:
                                ct_label = f"{b12+feeder_num-1}CT"
                            else:
                                ct_label = f"{b12+feeder_num}CT"
                    else:
                        if (bay_num - 1) % 3 == 0:     
                            ct_label = f"{b12+feeder_num+1}CT"
                        elif bay_num % 3 == 0:          
                            ct_label = f"{b12+feeder_num-1}CT"
                        else:
                            ct_label = f"{int(b12+feeder_num)}CT"

                    if current_type == "ICT":
                        iso_lbl2 = f"{b12+feeder_num}89T"
                        earth_lbl3 = f"{b12+feeder_num}89TE"
                    elif current_type == "Line_Bay":
                        iso_lbl2 = f"{b12+feeder_num}89L"
                        earth_lbl3 = f"{b12+feeder_num}89LE"
                    elif current_type == "Reactor":
                        iso_lbl2 = f"{b12+feeder_num}89R"
                        earth_lbl3 = f"{b12+feeder_num}89RE"
                    else:
                        iso_lbl2 = f"{b12+feeder_num}89C"
                        earth_lbl3 = f"{b12+feeder_num}89CE"

                    labels = {
                        "base_isolator" : f"{b12+feeder_num}89A",
                        "base_isolatorb" : f"{b12+feeder_num}89B",
                        "breaker_lbl" : f"{b12+feeder_num}52",
                        "earth_lbl1" : earth_label ,
                        "earth_lbl_BUS" : earth_label2 ,
                        "earth_lbl2" : f"{b12+feeder_num}89BE",
                        "iso_lbl2" : iso_lbl2,
                        "ct_lbl" : f"{int((b12+feeder_num))}CT",
                        "ct_lbl_middle" : ct_label ,
                        "wt_lbl" : f"{int((b12+feeder_num))}WT",
                        "cvt_lbl" : f"{int((b12+feeder_num))}CVT",
                        "la_lbl" : f"{int((b12+feeder_num))}LA",
                        "symbol_lbl" : f"{int((b12+feeder_num))}",
                        "earth_lbl3" : earth_lbl3,
                        "ict_lbl" : f"{int((b12+feeder_num))}ICT",
                        "rect_lbl" : f"{int((b12+feeder_num))}Reactor"
                    }
                    n=i+1
                    if (n - 3) % 3 == 0:
                        labels["base_isolator"], labels["base_isolatorb"],labels["earth_lbl1"],labels["earth_lbl2"] = (
                        labels["base_isolatorb"],
                        labels["base_isolator"],
                        labels["earth_lbl2"],
                        labels["earth_lbl1"],
                        )
                    return labels
                
                else: 
                    if (i + 1 == d6) or (i + 1 == d7):
                        earth_label = e7
                    else:
                        earth_label = b16

                    BCT_name = sheet["F6"].value
                    ACT_name = sheet["F7"].value
                    CT_name  = sheet["B22"].value

                    if ((bay_num % 3 == 1) or (bay_num % 3 == 0)):
                        pair_index = i - 2 if (bay_num % 3 == 0) else i + 2
                        pair_type = feeder_types[pair_index] if pair_index < len(feeder_types) else ""

                        if current_type in ["Line_Bay", "Future_Bay"] and pair_type in ["Line_Bay", "Future_Bay"]:
                            if (bay_num - 1) % 3 == 0:  
                                ct_label = f"{ACT_name}"
                            elif bay_num % 3 == 0:      
                                ct_label = f"{BCT_name}"
                            else:
                                ct_label = f"{CT_name}"
                        else:
                            ct_label = f"{CT_name}"
                    else:
                        ct_label = f"{CT_name}"
                        
                    labels = {
                        "base_isolator" : b14,
                        "base_isolatorb" : b15,
                        "breaker_lbl" : b17,
                        "earth_lbl1" : earth_label,
                        "earth_lbl_BUS" : e6 ,
                        "earth_lbl2" : b18,
                        "iso_lbl2" : b19,
                        "ct_lbl" : b22,
                        "ct_lbl_middle" : ct_label ,
                        "wt_lbl" : b23,
                        "cvt_lbl" : b24,
                        "la_lbl" : b25,
                        "symbol_lbl" : f"{int((b12+feeder_num))}",
                        "earth_lbl3" : b20,
                        "ict_lbl" : b26,
                        "rect_lbl" : b27
                    }
                    n=i+1
                    if (n - 3) % 3 == 0:
                        labels["base_isolator"], labels["base_isolatorb"],labels["earth_lbl1"],labels["earth_lbl2"] = (
                        labels["base_isolatorb"],
                        labels["base_isolator"],
                        labels["earth_lbl2"],
                        labels["earth_lbl1"],
                        )
                    return labels

            def common(ax, x_offset, y_offset, feeder_num, fs):
                L = get_labels(feeder_num)
                draw_isolator(ax, x_offset+.25, y_offset,L["base_isolator"], fs)
                earth_sh(ax, x_offset+.25, y_offset-.5, L["earth_lbl1"], fs)
                draw_breaker(ax, x_offset+0.25, y_offset-2.2, L["breaker_lbl"], fs)
                earth_sh(ax, x_offset+0.25, y_offset-3.6, L["earth_lbl2"], fs)
                draw_isolator(ax, x_offset+0.25, y_offset-4.6, L["base_isolatorb"], fs)
                draw_ct(ax, x_offset+0.25, y_offset-3.5, L["ct_lbl"], fs)
                ax.plot([x_offset+0.25, x_offset+0.25], [y_offset-3.2, y_offset-4.6+.2], color='red', linewidth=0.5)
                draw_name(ax, x_offset-0.5, y_offset-3,L["symbol_lbl"], fs)
                n=i+1
                if n == d6 and (n - 1) % 3 == 0:
                    earth_sh(ax, x_offset+.25, y_offset+1.1, L["earth_lbl_BUS"], fs)
                if n == d7 and (n - 3) % 3 == 0:
                    earth_sh(ax, x_offset+.25, y_offset-4.9, L["earth_lbl_BUS"], fs)

            def middle_common(ax, x_offset, y_offset, feeder_num, fs):
                L = get_labels(feeder_num)
                draw_isolator(ax, x_offset+.25, y_offset+1.1-.5,L["base_isolator"], fs)
                earth_sh(ax, x_offset+.25, y_offset+.6-.5, L["earth_lbl1"], fs)
                ax.plot([x_offset+0.25, x_offset+0.25], [y_offset-.15-.4, y_offset-.5-1.5], color='red', linewidth=0.5)
                draw_breaker(ax, x_offset+0.25, y_offset-2.2-.3-.5, L["breaker_lbl"], fs)
                earth_sh(ax, x_offset+0.25, y_offset-3.6-.1-.5, L["earth_lbl2"], fs)
                draw_isolator(ax, x_offset+0.25, y_offset-4.6-.1-.5, L["base_isolatorb"], fs)
                ax.plot([x_offset+0.25, x_offset+0.25], [y_offset-3.2-.8, y_offset-4.6-.6], color='red', linewidth=0.5)
                draw_name(ax, x_offset-0.5, y_offset-3-.6,L["symbol_lbl"], fs)

            def draw_feeder1(ax, x_offset, y_offset, feeder_num, fs):
                L = get_labels(feeder_num)
                n = i+1
                if (n - 1) % 3 == 0:
                    common(ax, x_offset, y_offset, feeder_num, fs)
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset+.9, y_offset+.4], color='red', linewidth=0.5)
                    ax.plot([x_offset+.25, x_offset+.75], [y_offset-5.8, y_offset-5.8], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset-5.8, y_offset+1.2], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset+2.2, y_offset+3.05], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset+3.25, y_offset+6.3], color='red', linewidth=0.5)
                    draw_isolator(ax, x_offset+0.75,y_offset+2 , L["iso_lbl2"], fs)
                    earth_sh(ax, x_offset+0.75, y_offset+3, L["earth_lbl3"], fs)
                    draw_wt(ax, x_offset+.75, y_offset+3.2, L["wt_lbl"], fs)
                    draw_cvt(ax, x_offset+.75, y_offset+5, L["cvt_lbl"], fs)
                    draw_la(ax, x_offset+.75, y_offset+5.2, L["la_lbl"], fs)
                    la_comp(ax, x_offset+.3, y_offset+5)
                    draw_symbol_upp(ax, x_offset+.75, y_offset+6.4, L["symbol_lbl"], fs)
                    draw_ct(ax, x_offset+0.25, y_offset-10.9, L["ct_lbl_middle"], fs)
                elif (n - 2) % 3 == 0:
                    middle_common(ax, x_offset, y_offset, feeder_num, fs)
                elif (n - 3) % 3 == 0:
                    common(ax, x_offset, y_offset, feeder_num, fs)
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset-5.8, y_offset-6.4], color='red', linewidth=0.5)
                    ax.plot([x_offset+.25, x_offset+.75], [y_offset+.5, y_offset+.5], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset+.5, y_offset-7], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset-8, y_offset-8.55], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset-8.75, y_offset-11], color='red', linewidth=0.5)
                    draw_isolator(ax, x_offset+0.75, y_offset-7, L["iso_lbl2"], fs)
                    earth_sh(ax, x_offset+0.75, y_offset-7.3, L["earth_lbl3"], fs)
                    draw_wt(ax, x_offset+.75, y_offset-7.6-1, L["wt_lbl"], fs)
                    draw_cvt(ax, x_offset+.75, y_offset-7.7-1, L["cvt_lbl"], fs)
                    draw_la(ax, x_offset+.75, y_offset-9-1, L["la_lbl"], fs)
                    la_comp(ax, x_offset+.3, y_offset-9.2-1)
                    draw_symbol(ax, x_offset+.75, y_offset-10.1-1, L["symbol_lbl"], fs)
                    draw_ct(ax, x_offset+0.25, y_offset+5, L["ct_lbl_middle"], fs)
                else:
                    print("Invalid")

            def draw_feeder2(ax, x_offset, y_offset, feeder_num, fs):
                L = get_labels(feeder_num)
                n = i+1
                if (n - 1) % 3 == 0:
                    common(ax, x_offset, y_offset, feeder_num, fs)
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset+.9, y_offset+.4], color='red', linewidth=0.5)
                    ax.plot([x_offset+.25, x_offset+.75], [y_offset-5.8, y_offset-5.8], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset-5.8, y_offset+1.2], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset+2.2, y_offset+4.2], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset+5.4, y_offset+6.3], color='red', linewidth=0.5)
                    draw_isolator(ax, x_offset+0.75,y_offset+2 , L["iso_lbl2"], fs)
                    earth_sh(ax, x_offset+0.75, y_offset+3, L["earth_lbl3"], fs)
                    draw_la(ax, x_offset+.75, y_offset+3.6, L["la_lbl"], fs)
                    la_comp(ax, x_offset+.3, y_offset+3.4)
                    draw_ict_upp(ax,x_offset+0.75, y_offset+5.2, L["ict_lbl"], fs)
                    draw_symbol_upp(ax, x_offset+.75, y_offset+6.4, L["symbol_lbl"], fs)
                elif (n - 2) % 3 == 0:
                    middle_common(ax, x_offset, y_offset, feeder_num, fs)
                elif (n - 3) % 3 == 0:
                    common(ax, x_offset, y_offset, feeder_num, fs)
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset-5.8, y_offset-6.4], color='red', linewidth=0.5)
                    ax.plot([x_offset+.25, x_offset+.75], [y_offset+.5, y_offset+.5], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset+.5, y_offset-7], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset-8, y_offset-9.8], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset-11, y_offset-11.5], color='red', linewidth=0.5)
                    draw_isolator(ax, x_offset+0.75, y_offset-7, L["iso_lbl2"], fs)
                    earth_sh(ax, x_offset+0.75, y_offset-7.3, L["earth_lbl3"], fs)
                    draw_ict(ax,x_offset+0.75, y_offset-10, L["ict_lbl"], fs)
                    draw_la(ax, x_offset+.75, y_offset-8-1, L["la_lbl"], fs)
                    la_comp(ax, x_offset+.3, y_offset-8.2-1)
                    draw_symbol(ax, x_offset+.75, y_offset-11.1-.5, L["symbol_lbl"], fs)
                else:
                    print("Invalid")

            def draw_feeder3(ax, x_offset, y_offset, feeder_num, fs):
                L = get_labels(feeder_num)
                n = i+1
                if (n - 1) % 3 == 0:
                    common(ax, x_offset, y_offset, feeder_num, fs)
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset+.9, y_offset+.4], color='red', linewidth=0.5)
                    ax.plot([x_offset+.25, x_offset+.75], [y_offset-5.8, y_offset-5.8], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset-5.8, y_offset+1.2], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset+2.2, y_offset+4.2], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset+5.35, y_offset+6.3], color='red', linewidth=0.5)
                    draw_isolator(ax, x_offset+0.75,y_offset+2 , L["iso_lbl2"], fs)
                    earth_sh(ax, x_offset+0.75, y_offset+3, L["earth_lbl3"], fs)
                    draw_la(ax, x_offset+.75, y_offset+3.6, L["la_lbl"], fs)
                    la_comp(ax, x_offset+.3, y_offset+3.4)
                    draw_reacter(ax,x_offset+0.75, y_offset+5.2, L["rect_lbl"], fs)
                    draw_earth_symbol_upp(ax, x_offset+.75, y_offset+6.6, L["symbol_lbl"], fs)
                elif (n - 2) % 3 == 0:
                    middle_common(ax, x_offset, y_offset, feeder_num, fs)
                elif (n - 3) % 3 == 0:
                    common(ax, x_offset, y_offset, feeder_num, fs)
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset-5.8, y_offset-6.4], color='red', linewidth=0.5)
                    ax.plot([x_offset+.25, x_offset+.75], [y_offset+.5, y_offset+.5], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset+.5, y_offset-7], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset-8, y_offset-9.85], color='red', linewidth=0.5)
                    ax.plot([x_offset+.75, x_offset+.75], [y_offset-11, y_offset-11.5], color='red', linewidth=0.5)
                    draw_isolator(ax, x_offset+0.75, y_offset-7, L["iso_lbl2"], fs)
                    earth_sh(ax, x_offset+0.75, y_offset-7.3, L["earth_lbl3"], fs)
                    draw_reacter(ax,x_offset+0.75, y_offset-10, L["rect_lbl"], fs)
                    draw_la(ax, x_offset+.75, y_offset-8-1, L["la_lbl"], fs)
                    la_comp(ax, x_offset+.3, y_offset-8.2-1)
                    draw_earth_symbol(ax, x_offset+.75, y_offset-11.1-.5, L["symbol_lbl"], fs)
                else:
                    print("Invalid")

            def draw_feeder4(ax, x_offset, y_offset, feeder_num, fs):
                L = get_labels(feeder_num)
                n = i+1
                if (n - 1) % 3 == 0:
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset+.9, y_offset+.4], color='red', linewidth=0.5)
                    draw_isolator(ax, x_offset+.25, y_offset,L["base_isolator"], fs)
                    earth_sh(ax, x_offset+.25, y_offset-.5, L["earth_lbl1"], fs)
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset-1, y_offset-5.95], color='red', linewidth=0.5)
                    draw_name(ax, x_offset-0.5, y_offset-3,L["symbol_lbl"], fs)
                    draw_ct(ax, x_offset+0.25, y_offset-10.9, L["ct_lbl_middle"], fs)
                elif (n - 2) % 3 == 0:
                    middle_common(ax, x_offset, y_offset, feeder_num, fs)
                elif (n - 3) % 3 == 0:
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset-5.8, y_offset-6.4], color='red', linewidth=0.5)
                    earth_sh(ax, x_offset+0.25, y_offset-3.6, L["earth_lbl2"], fs)
                    draw_isolator(ax, x_offset+0.25, y_offset-4.6, L["base_isolatorb"], fs)
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset+.45, y_offset-4.6], color='red', linewidth=0.5)
                    draw_name(ax, x_offset-0.5, y_offset-3,L["symbol_lbl"], fs)
                    draw_ct(ax, x_offset+0.25, y_offset+5, L["ct_lbl_middle"], fs)
                else:
                    print("Invalid")

            def draw_feeder5(ax, x_offset, y_offset, feeder_num, fs):
                L = get_labels(feeder_num)
                n = i+1
                if (n - 1) % 3 == 0:
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset+.9, y_offset-6], color='red', linewidth=0.5)
                    draw_name(ax, x_offset-0.5, y_offset-3,L["symbol_lbl"], fs)
                elif (n - 2) % 3 == 0:
                    middle_common(ax, x_offset, y_offset, feeder_num, fs)
                elif (n - 3) % 3 == 0:
                    ax.plot([x_offset+0.25, x_offset+0.25], [y_offset+.5, y_offset-6.4], color='red', linewidth=0.5)
                    draw_name(ax, x_offset-0.5, y_offset-3,L["symbol_lbl"], fs)
                else:
                    print("Invalid")

            fig_width = max(12, num_feeders/4 *6)
            fig, ax = plt.subplots(figsize=(fig_width, 18))
            x_start = 5
            y_start = 19.8
            gap = 3
            fontsize = get_fontsize(num_feeders)

            feeders_per_column = 3
            y_gap = 6.2+.6  
            x_gap = gap  

            ax.plot([1, num_feeders*2], [20.7,20.7], color='blue', linewidth=0.5)
            ax.plot([1, num_feeders*2], [-.2,-.2], color='green', linewidth=0.5)

            for i in range(num_feeders):
                f_type = feeder_types[i] if i < len(feeder_types) else '1'
                f_name = str(feeder_names[i]) if i < len(feeder_names) else ''

                col = i // feeders_per_column       
                row = i % feeders_per_column        
                x_pos = x_start + col * x_gap
                y_pos = y_start - row * y_gap       
                feeder_label = i + 1

                if f_type == 'Line_Bay': draw_feeder1(ax, x_pos, y_pos, feeder_label, fontsize)
                elif f_type == 'ICT': draw_feeder2(ax, x_pos, y_pos, feeder_label, fontsize)
                elif f_type == 'Reactor': draw_feeder3(ax, x_pos, y_pos, feeder_label, fontsize)
                elif f_type == 'Future_Bay': draw_feeder4(ax, x_pos, y_pos, feeder_label, fontsize)
                elif f_type == 'No_bay': draw_feeder5(ax, x_pos, y_pos, feeder_label, fontsize)
                else: draw_feeder1(ax, x_pos, y_pos, feeder_label, fontsize)

                n = i + 1
                words = f_name.split()
                lines, current_line = [], ""
                for word in words:
                    if len(current_line + " " + word) <= 24:
                        if current_line: current_line += " " + word
                        else: current_line = word
                    else:
                        lines.append(current_line)
                        current_line = word
                if current_line: lines.append(current_line)
                wrapped_text = "\n".join(lines)

                label_text = f"{b12 + feeder_label}"
                full_text = f"{label_text}\n{wrapped_text}"
                
                if n % 3 == 1:      # 1, 4, 7, 10 ...
                    y_pos = y_start + 8.5
                elif n % 3 == 2:    # 2, 5, 8, 11 ...
                    y_pos = None
                else:                # 3, 6, 9, 12 ...
                    y_pos = y_start-26
                
                if y_pos is not None:
                    ax.text(x_pos+.75, y_pos, full_text, ha='center', va='top', fontsize=fontsize+1)
            
            center_x = (x_start + (num_feeders/3)*gap)/2
            ax.text(1.5, 21.1, f"{b12} KV_BUS1", fontsize=fontsize+4, va='center')
            ax.text(1.5, .1, f"{b12} KV_BUS2", fontsize=fontsize+4, va='center')
            ax.text(center_x, 32, 'POWERGRID CORPORATION OF INDIA LTD', fontsize=fontsize+25, va='center', ha='center')  
            ax.text(center_x, 30, b9, fontsize=fontsize+15, va='center', ha='center')  

            num_columns = (num_feeders + feeders_per_column - 1) // feeders_per_column
            ax.set_xlim(0, x_start + num_feeders/3 * gap+ x_start)
            ax.set_ylim(-10,35)  
            ax.axis('off')

            img_path = os.path.join(os.path.dirname(file_path), "temp_sld.png")
            fig.savefig(img_path, bbox_inches='tight', dpi=300)

            if "Feeder Diagram" in wb.sheetnames:
                std = wb["Feeder Diagram"]
                wb.remove(std)

            ws_new = wb.create_sheet("Feeder Diagram")
            img = XLImage(img_path)
            img.width = num_feeders*120/3*2
            img.height = 250*2.5*2
            ws_new.add_image(img, "A1")

            wb.save(file_path)
            wb.close()

            # Generate Outputs
            st.pyplot(fig)
            pdf_io = io.BytesIO()
            fig.savefig(pdf_io, format='pdf', bbox_inches='tight')
            dxf_doc = export_ax_to_dxf(ax)
            dxf_io = io.StringIO()
            dxf_doc.write(dxf_io)

            st.success("Generation Complete!")
            col1, col2, col3 = st.columns(3)
            with col1:
                with open(file_path, "rb") as f:
                    st.download_button(label="📥 Download Excel", data=f, file_name="SLD_Integrated.xlsx")
            with col2:
                st.download_button(label="📥 Download PDF", data=pdf_io.getvalue(), file_name="Substation_SLD.pdf", mime="application/pdf")
            with col3:
                st.download_button(label="📥 Download AutoCAD DXF", data=dxf_io.getvalue(), file_name="Substation_SLD.dxf", mime="application/dxf")
            
            os.remove(img_path)

    except Exception as e:
        import traceback
        st.error(f"Error executing file: {str(e)}")
        st.code(traceback.format_exc())
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)
